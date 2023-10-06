from openpyxl import load_workbook


def workbook_excel(filename: str):
    workbook = load_workbook(filename=filename)
    return workbook


# path_ods = os.path.join(os.getcwd(), 'data_xlsx', 'file.xlsx')
path_ods = '/Users/mac/python_projects/python_sandbox/run_python_in_excel/data_xlsx/file.xlsx'
wb = workbook_excel(filename=path_ods)
ws = wb.worksheets[0]
cell_value = ws['A1'].value
new_val_cell = cell_value + 1
ws['A1'] = new_val_cell
wb.save(path_ods)


with open('/Users/mac/python_projects/python_sandbox/run_python_in_excel/check_macros.log', 'w') as f:
    f.write('OK')

